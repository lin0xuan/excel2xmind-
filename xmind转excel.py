from typing import List, Any
from openpyxl import Workbook
from xmindparser import xmind_to_dict


def resolve_path(dict_, lists, title):
    """
    通过递归取出每个主分支下的所有小分支并将其作为一个列表
    :param dict_:
    :param lists:
    :param title:
    :return:
    """
    # 去除title首尾空格
    title = title.strip()
    # 若title为空，则直接取value
    if len(title) == 0:
        concat_title = dict_["title"].strip()
    else:
        concat_title = title + "\t" + dict_["title"].strip()
    if not dict_.__contains__("topics"):
        lists.append(concat_title)
    else:
        for d in dict_["topics"]:
            resolve_path(d, lists, concat_title)


def xmind_to_excel(list_, excel_path):
    wb = Workbook()
    # 生成单sheet的Excel文件，sheet名自取
    sheet = wb.active
    sheet.title = "XX模块"

    # 第一行固定的表头标题
    row_header = ["序号", "模块", "功能点"]
    for i in range(len(row_header)):
        sheet.cell(row=1, column=i+1, value=row_header[i])

    # 增量索引
    index = 0

    for h in range(len(list_)):
        lists: List[Any] = []
        resolve_path(list_[h], lists, "")

        for j in range(len(lists)):
            # 将主分支下的小分支构成列表
            lists[j] = lists[j].split('\t')

            for n in range(len(lists[j])):
                # 生成第一列的序号
                sheet.cell(row=j + index + 2, column=1, value=j + index + 1)
                sheet.cell(row=j + index + 2, column=n + 2, value=lists[j][n])
                # 自定义内容标题
                if n >= 2:
                    sheet.cell(row=1, column=n + 2, value=f"自定义{n - 1}")

        # 遍历完lists并给增量索引赋值
        if lists:  # 确保列表不为空
            index += len(lists)

    wb.save(excel_path)


def run(xmind_path):
    # 将XMind转化成字典
    xmind_dict = xmind_to_dict(xmind_path)
    # Excel文件与XMind文件保存在同一目录下
    excel_name = xmind_path.split('\\')[-1].split(".")[0] + '.xlsx'  # openpyxl支持xlsx格式
    excel_path = "\\".join(xmind_path.split('\\')[:-1]) + "\\" + excel_name
    print(f"生成的Excel路径: {excel_path}")
    # 处理XMind数据并生成Excel
    xmind_to_excel(xmind_dict[0]['topic']['topics'], excel_path)

if __name__ == '__main__':
    xmind_path_ = r"C:\Users\Ivan8\Desktop\笃威尔\开源情报\全产业链体系\全体系xmind\产业链标签体系.xmind"
    run(xmind_path_)
