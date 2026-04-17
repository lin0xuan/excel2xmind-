import json
import uuid
import zipfile
from io import BytesIO
import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

# ==============================
# 树形节点结构
# ==============================
class TreeNode:
    def __init__(self, text):
        self.id = str(uuid.uuid4())
        self.text = text
        self.children = []

# ==============================
# 1. 读取 Excel + 处理合并单元格（新版核心功能）
# ==============================
def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # 初始化空数据表格
    data = []
    for r in range(max_row):
        data.append(["" for _ in range(max_col)])

    # 读取所有单元格基础值
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            data[row_idx - 1][col_idx - 1] = str(cell.value).strip() if cell.value is not None else ""

    # ======================
    # 处理合并单元格（和新版TS逻辑完全一致）
    # ======================
    if ws.merged_cells:
        for merged_range in ws.merged_cells.ranges:
            # 获取合并区域的起止坐标
            min_row, min_col, max_row, max_col = (
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col
            )

            # 获取合并区域左上角单元格的值（主值）
            top_left_value = ws.cell(min_row, min_col).value
            fill_value = str(top_left_value).strip() if top_left_value is not None else ""

            # 给合并区域内所有单元格赋值
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    data[r - 1][c - 1] = fill_value

    return data

# ==============================
# 2. Excel → 树形结构（和新版TS逻辑完全一致）
# ==============================
def excel_to_tree(data):
    root = TreeNode("思维导图")

    for row in data:
        current_parent = root

        for cell_text in row:
            cell_text = str(cell_text or "").strip()
            if not cell_text:
                continue

            # 查找同级是否已存在相同文本节点
            existing_node = None
            for child in current_parent.children:
                if child.text == cell_text:
                    existing_node = child
                    break

            if not existing_node:
                existing_node = TreeNode(cell_text)
                current_parent.children.append(existing_node)

            current_parent = existing_node

    # 清理空的 children
    def clean_empty(node):
        if not node.children:
            node.children = None
        else:
            for child in node.children:
                clean_empty(child)

    clean_empty(root)
    return root

# ==============================
# 3. 转换为标准 XMind 格式
# ==============================
def tree_to_xmind_content(root):
    def convert_node(node):
        res = {
            "id": node.id,
            "title": node.text,
            "structureClass": "org.xmind.ui.map.unbalanced"
        }

        if node.children and len(node.children) > 0:
            res["children"] = {
                "attached": [convert_node(child) for child in node.children]
            }
        return res

    return [
        {
            "id": str(uuid.uuid4()),
            "title": "Sheet1",
            "rootTopic": convert_node(root)
        }
    ]

# ==============================
# 4. 生成 XMind ZIP 文件
# ==============================
def generate_xmind_zip(content):
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))

        manifest = {
            "file-entries": {
                "content.json": {},
                "metadata.json": {}
            }
        }
        zf.writestr("manifest.json", json.dumps(manifest, indent=2))

        metadata = {
            "creator": {
                "name": "Excel to XMind Converter",
                "version": "1.0.0"
            },
            "created": "2026-01-01T00:00:00.000Z",
            "version": "3.7.0"
        }
        zf.writestr("metadata.json", json.dumps(metadata, indent=2))

    zip_buffer.seek(0)
    return zip_buffer

# ==============================
# 主函数
# ==============================
def excel2xmind(excel_path, output_path):
    print("📖 读取 Excel 文件（已支持合并单元格）...")
    data = read_excel(excel_path)

    print("🌲 构建思维导图结构...")
    tree = excel_to_tree(data)

    print("📝 生成 XMind 内容...")
    xmind_content = tree_to_xmind_content(tree)

    print("📦 打包生成 .xmind 文件...")
    xmind_zip = generate_xmind_zip(xmind_content)

    with open(output_path, "wb") as f:
        f.write(xmind_zip.getvalue())

    print(f"✅ 转换完成！文件已保存到：{output_path}")

# ==============================
# 直接运行
# ==============================
if __name__ == "__main__":
    # 输入输出路径（可自行修改）
    INPUT_EXCEL = r"C:\Users\Ivan8\Desktop\temp_体系转xmind.xlsx"
    OUTPUT_XMIND = r"C:\Users\Ivan8\Desktop\output.xmind"
    
    excel2xmind(INPUT_EXCEL, OUTPUT_XMIND)